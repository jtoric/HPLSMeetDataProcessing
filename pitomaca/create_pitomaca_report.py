import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import math

def calculate_ipf_gl_points(bodyweight, total, sex, event):
    """Calculate IPF GL Points using official IPF formula: 100/(A-B*exp(-C*bodyweight))"""
    if pd.isna(bodyweight) or pd.isna(total) or total == 0:
        return 0
    
    # Official IPF GL Coefficients from IPF_GL_Coefficients-2020.pdf
    if event == 'SBD':  # Classic Powerlifting
        if sex == 'M':  # Men's Classic Powerlifting
            A = 1199.72839
            B = 1025.18162
            C = 0.00921
        else:  # Women's Classic Powerlifting
            A = 610.32796
            B = 1045.59282
            C = 0.03048
    else:  # Bench Only (Classic Bench Press)
        if sex == 'M':  # Men's Classic Bench Press
            A = 320.98041
            B = 281.40258
            C = 0.01008
        else:  # Women's Classic Bench Press
            A = 142.40398
            B = 442.52671
            C = 0.04724
    
    try:
        # IPF GL Points formula: 100/(A-B*exp(-C*bodyweight))
        ipf_gl_points = 100 / (A - B * math.exp(-C * bodyweight)) * total
        
        return round(ipf_gl_points, 2)
    except:
        return 0

def load_club_mapping():
    """Load club and birth year mapping from klubovi.csv"""
    club_mapping = {}
    birth_year_mapping = {}
    
    try:
        df = pd.read_csv('pitomaca/klubovi.csv', header=None)
        
        for _, row in df.iterrows():
            # Find rows that have name data (check columns 3 and 4 for IME and PREZIME)
            try:
                ime = row[3]
                prezime = row[4]
                godiste = row[5]
                klub = row[6]
                
                if pd.notna(ime) and pd.notna(prezime) and str(ime).strip() != '' and str(prezime).strip() != '':
                    # Skip header rows
                    if str(ime).upper() == 'IME' or str(prezime).upper() == 'PREZIME':
                        continue
                    
                    full_name = f"{str(ime).strip()} {str(prezime).strip()}"
                    
                    if pd.notna(klub) and str(klub).strip() != '':
                        club_mapping[full_name] = str(klub).strip()
                    
                    if pd.notna(godiste):
                        try:
                            year = int(float(godiste))
                            birth_year_mapping[full_name] = year
                        except:
                            pass
            except:
                continue
        
        print(f"Učitano {len(club_mapping)} mapiranja klubova")
        print(f"Učitano {len(birth_year_mapping)} mapiranja godina rođenja")
        
    except Exception as e:
        print(f"Greška pri učitavanju klubovi.csv: {e}")
    
    return club_mapping, birth_year_mapping

def get_division_type(division_name):
    """Extract division type from full division name"""
    if 'Masters 2' in division_name or 'Master II' in division_name:
        return 'Master II'
    elif 'Masters 1' in division_name or 'Master I' in division_name:
        return 'Master I'
    elif 'Sub-Junior' in division_name or 'Sub-Juniors' in division_name:
        return 'Sub-Junior'
    elif 'Junior' in division_name and 'Sub' not in division_name:
        return 'Junior'
    elif 'Open' in division_name:
        return 'Open'
    else:
        return 'Open'  # Default

def translate_column_headers(columns):
    """Translate English column headers to Croatian"""
    translation_map = {
        'Place': 'Plasman',
        'Name': 'Ime i prezime',
        'Club': 'Klub',
        'Sex': 'Spol',
        'BirthYear': 'Godina rođenja',
        'Division': 'Kategorija',
        'BodyweightKg': 'Tjelesna masa (kg)',
        'WeightClassKg': 'Težinska kategorija (kg)',
        'Squat1Kg': 'Čučanj 1 (kg)',
        'Squat2Kg': 'Čučanj 2 (kg)',
        'Squat3Kg': 'Čučanj 3 (kg)',
        'Best3SquatKg': 'Najbolji čučanj (kg)',
        'Bench1Kg': 'Potisak s klupe 1 (kg)',
        'Bench2Kg': 'Potisak s klupe 2 (kg)',
        'Bench3Kg': 'Potisak s klupe 3 (kg)',
        'Best3BenchKg': 'Najbolji potisak s klupe (kg)',
        'Deadlift1Kg': 'Mrtvo dizanje 1 (kg)',
        'Deadlift2Kg': 'Mrtvo dizanje 2 (kg)',
        'Deadlift3Kg': 'Mrtvo dizanje 3 (kg)',
        'Best3DeadliftKg': 'Najbolje mrtvo dizanje (kg)',
        'TotalKg': 'Ukupno (kg)',
        'Points': 'GL Bodovi',
        'Event': 'Disciplina'
    }
    
    return [translation_map.get(col, col) for col in columns]

def translate_division_name(division_name, sex):
    """Translate English division names to Croatian with gender"""
    gender_prefix = "Muški " if sex == 'M' else "Ženski "
    
    # Translate division types
    if 'Masters 2' in division_name or 'Master II' in division_name:
        return gender_prefix + "Veterani 2"
    elif 'Masters 1' in division_name or 'Master I' in division_name:
        return gender_prefix + "Veterani 1"
    elif 'Sub-Junior' in division_name or 'Sub-Juniors' in division_name:
        return gender_prefix + "Kadeti"
    elif 'Junior' in division_name and 'Sub' not in division_name:
        return gender_prefix + "Juniori"
    elif 'Open' in division_name:
        return gender_prefix + "Seniori"
    else:
        return gender_prefix + "Seniori"  # Default

def translate_division_type(division_type):
    """Translate division type for headers"""
    translations = {
        'Sub-Junior': 'KADETI',
        'Junior': 'JUNIORI',
        'Open': 'SENIORI',
        'Master I': 'VETERANI 1',
        'Master II': 'VETERANI 2', 
        'Master III': 'VETERANI 3',
        'Master IV': 'VETERANI 4'
    }
    return translations.get(division_type, division_type.upper())

def auto_fit_columns(worksheet):
    """Auto-fit all column widths in the worksheet"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        
        for cell in column:
            try:
                if hasattr(cell, 'column_letter') and cell.column_letter:
                    if column_letter is None:
                        column_letter = cell.column_letter
                    
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            except:
                pass
        
        if column_letter is not None:
            adjusted_width = min(max_length + 3, 30)
            if adjusted_width < 8:
                adjusted_width = 8
            worksheet.column_dimensions[column_letter].width = adjusted_width

def process_pitomaca_data():
    """Process Pitomača CSV data into standardized format"""
    
    # Load club and birth year mappings
    club_mapping, birth_year_mapping = load_club_mapping()
    
    # Read the CSV
    df = pd.read_csv('pitomaca/hpls_2504.csv')
    
    print(f"Ukupno redaka u CSV: {len(df)}")
    
    # Filter out non-valid results (NS = No Show, DQ = Disqualified)
    df_valid = df[~df['Place'].isin(['NS', 'DQ', ''])]
    df_valid = df_valid.dropna(subset=['Place'])
    
    # Convert Place to numeric where possible
    df_valid = df_valid.copy()
    df_valid['PlaceNum'] = pd.to_numeric(df_valid['Place'], errors='coerce')
    df_valid = df_valid[df_valid['PlaceNum'].notna()]
    
    print(f"Važeći rezultati: {len(df_valid)}")
    
    # Create standardized output
    output_data = []
    
    for _, row in df_valid.iterrows():
        # Create full division name combining Sex, Division and Event
        sex = row['Sex']
        division = row['Division']
        event = row['Event']
        name = row['Name']
        
        # Create Division name like "Men's Raw Open" or "Women's Raw Sub-Juniors"
        gender = "Men's" if sex == 'M' else "Women's"
        event_suffix = " Bench Only" if event == 'B' else ""
        full_division = f"{gender} Raw {division}{event_suffix}"
        
        # Calculate GL points (use existing Goodlift if available, otherwise calculate)
        gl_points = row['Goodlift'] if pd.notna(row['Goodlift']) and row['Goodlift'] > 0 else calculate_ipf_gl_points(
            bodyweight=pd.to_numeric(row['BodyweightKg'], errors='coerce'),
            total=pd.to_numeric(row['TotalKg'], errors='coerce') if event == 'SBD' else pd.to_numeric(row['Best3BenchKg'], errors='coerce'),
            sex=sex,
            event=event
        )
        
        # Get total - for bench only it's the best bench
        if event == 'B':
            total = row['Best3BenchKg'] if pd.notna(row['Best3BenchKg']) else 0
        else:
            total = row['TotalKg'] if pd.notna(row['TotalKg']) else 0
        
        # Get club from mapping
        club = club_mapping.get(name, '')
        
        # Get birth year from mapping
        birth_year = birth_year_mapping.get(name, '')
        
        output_data.append({
            'Place': int(row['PlaceNum']),
            'Name': name,
            'Club': club,
            'Sex': sex,
            'BirthYear': birth_year,
            'Division': full_division,
            'BodyweightKg': row['BodyweightKg'],
            'WeightClassKg': row['WeightClassKg'],
            'Squat1Kg': row['Squat1Kg'] if pd.notna(row['Squat1Kg']) else '',
            'Squat2Kg': row['Squat2Kg'] if pd.notna(row['Squat2Kg']) else '',
            'Squat3Kg': row['Squat3Kg'] if pd.notna(row['Squat3Kg']) else '',
            'Best3SquatKg': row['Best3SquatKg'] if pd.notna(row['Best3SquatKg']) else '',
            'Bench1Kg': row['Bench1Kg'] if pd.notna(row['Bench1Kg']) else '',
            'Bench2Kg': row['Bench2Kg'] if pd.notna(row['Bench2Kg']) else '',
            'Bench3Kg': row['Bench3Kg'] if pd.notna(row['Bench3Kg']) else '',
            'Best3BenchKg': row['Best3BenchKg'] if pd.notna(row['Best3BenchKg']) else '',
            'Deadlift1Kg': row['Deadlift1Kg'] if pd.notna(row['Deadlift1Kg']) else '',
            'Deadlift2Kg': row['Deadlift2Kg'] if pd.notna(row['Deadlift2Kg']) else '',
            'Deadlift3Kg': row['Deadlift3Kg'] if pd.notna(row['Deadlift3Kg']) else '',
            'Best3DeadliftKg': row['Best3DeadliftKg'] if pd.notna(row['Best3DeadliftKg']) else '',
            'TotalKg': total,
            'Points': gl_points,
            'Event': event
        })
    
    result_df = pd.DataFrame(output_data)
    
    # Check for missing club or birth year data - raise error if found
    missing_data = []
    for _, row in result_df.iterrows():
        name = row['Name']
        missing = []
        if row['Club'] == '':
            missing.append('klub')
        if row['BirthYear'] == '':
            missing.append('godina rođenja')
        if missing:
            missing_data.append(f"  - {name}: nedostaje {', '.join(missing)}")
    
    if missing_data:
        error_msg = f"\n❌ GREŠKA: Nedostaju podaci za {len(missing_data)} natjecatelja:\n"
        error_msg += "\n".join(missing_data)
        error_msg += "\n\nMolimo ažurirajte klubovi.csv datoteku s podacima koji nedostaju."
        raise ValueError(error_msg)
    
    print(f"✓ Pronađeni svi podaci (klub i godina rođenja) za {len(result_df)} natjecatelja")
    
    return result_df

def sort_by_categories(df):
    """Sort dataframe by division order, then weight class, then place"""
    
    division_order = {
        'Sub-Junior': 1,
        'Junior': 2, 
        'Open': 3,
        'Master I': 4,
        'Master II': 5,
        'Master III': 6,
        'Master IV': 7
    }
    
    df = df.copy()
    df['DivisionOrder'] = df['Division'].apply(lambda x: division_order.get(get_division_type(x), 3))
    
    def weight_sort_key(weight_class_str):
        weight_str = str(weight_class_str)
        try:
            if '+' in weight_str:
                base_weight = float(weight_str.replace('+', ''))
                return base_weight + 0.5
            else:
                return float(weight_str)
        except ValueError:
            return 9999.0
    
    df['WeightClassKg_num'] = df['WeightClassKg'].apply(weight_sort_key)
    df['Place_num'] = pd.to_numeric(df['Place'], errors='coerce')
    
    df_sorted = df.sort_values(['DivisionOrder', 'WeightClassKg_num', 'Place_num'])
    df_sorted = df_sorted.drop(['DivisionOrder', 'WeightClassKg_num', 'Place_num'], axis=1)
    
    return df_sorted

def generate_club_rankings(df):
    """Generate club ranking CSV files based on top 5 lifters per club"""
    
    categories = [
        (df[(df['Sex'] == 'M') & (df['Event'] == 'SBD')], 'Male_Powerlifting_Ranking.csv', 'Muški Powerlifting'),
        (df[(df['Sex'] == 'F') & (df['Event'] == 'SBD')], 'Female_Powerlifting_Ranking.csv', 'Ženski Powerlifting'),
        (df[(df['Sex'] == 'M') & (df['Event'] == 'B')], 'Male_Bench_Only_Ranking.csv', 'Muški Potisak s klupe'),
        (df[(df['Sex'] == 'F') & (df['Event'] == 'B')], 'Female_Bench_Only_Ranking.csv', 'Ženski Potisak s klupe')
    ]
    
    rankings = {}
    
    for data, filename, category_name in categories:
        print(f"\nGeneriranje rangiranja za {category_name}...")
        
        # Filter out empty clubs
        data_with_clubs = data[data['Club'] != ''].copy()
        
        if data_with_clubs.empty:
            print(f"  Nema podataka o klubovima za {category_name}")
            empty_df = pd.DataFrame(columns=['Place', 'Club', 'Points'])
            empty_df.to_csv(filename, index=False)
            rankings[filename] = empty_df
            continue
        
        # For each club, take only the top 5 lifters by Points
        def get_top_5_points(group):
            top_5 = group.nlargest(5, 'Points')
            return top_5['Points'].sum()
        
        club_totals = data_with_clubs.groupby('Club').apply(get_top_5_points, include_groups=False).reset_index()
        club_totals.columns = ['Club', 'Points']
        
        # Sort by total points (descending) and add ranking
        club_totals = club_totals.sort_values('Points', ascending=False).reset_index(drop=True)
        club_totals['Place'] = range(1, len(club_totals) + 1)
        
        # Reorder columns
        club_rankings = club_totals[['Place', 'Club', 'Points']].copy()
        club_rankings['Points'] = club_rankings['Points'].round(2)
        
        # Save the ranking
        club_rankings.to_csv(filename, index=False)
        rankings[filename] = club_rankings
        
        print(f"  Generirano {len(club_rankings)} klubova (top 5 liftera po klubu)")
        
        # Show top 3 clubs
        if len(club_rankings) > 0:
            print("  Top 3 klubova:")
            for _, row in club_rankings.head(3).iterrows():
                print(f"    {int(row['Place'])}. {row['Club']} - {row['Points']:.2f} bodova")
    
    return rankings

def create_formatted_sheet(worksheet, data, header_font, header_fill, header_alignment, data_font, data_alignment, border):
    """Create a formatted sheet with separate tables for each division/weight class combination"""
    
    if len(data) == 0:
        return
    
    current_row = 1
    division_order = ['Sub-Junior', 'Junior', 'Open', 'Master I', 'Master II', 'Master III', 'Master IV']
    
    data_copy = data.copy()
    data_copy['DivisionType'] = data_copy['Division'].apply(get_division_type)
    
    def weight_sort_key(weight_class_str):
        weight_str = str(weight_class_str)
        try:
            if '+' in weight_str:
                base_weight = float(weight_str.replace('+', ''))
                return base_weight + 0.5
            else:
                return float(weight_str)
        except ValueError:
            return 9999.0
    
    data_copy['WeightSortKey'] = data_copy['WeightClassKg'].apply(weight_sort_key)
    data_copy['PlaceNumeric'] = pd.to_numeric(data_copy['Place'], errors='coerce')
    
    division_order_map = {div: i for i, div in enumerate(division_order)}
    data_copy['DivisionOrder'] = data_copy['DivisionType'].map(division_order_map).fillna(999)
    data_sorted = data_copy.sort_values(['DivisionOrder', 'WeightSortKey', 'PlaceNumeric'])
    
    # Get original column names (excluding helper columns, Division and WeightClassKg since they're shown in headers)
    original_columns = [col for col in data.columns if col not in ['Division', 'WeightClassKg', 'Sex']]
    
    # Get unique division/weight class combinations in sorted order
    unique_combinations = []
    seen = set()
    for _, row in data_sorted.iterrows():
        combo = (row['Division'], row['WeightClassKg'])
        if combo not in seen:
            unique_combinations.append(combo)
            seen.add(combo)
    
    current_division_type = None
    
    for division, weight_class in unique_combinations:
        group_data = data_sorted[(data_sorted['Division'] == division) & 
                                (data_sorted['WeightClassKg'] == weight_class)]
        
        division_type = get_division_type(division)
        
        if current_division_type != division_type:
            if current_division_type is not None:
                current_row += 2
            
            translated_division_type = translate_division_type(division_type)
            division_header = f"═══ {translated_division_type} KATEGORIJA ═══"
            cell = worksheet.cell(row=current_row, column=1, value=division_header)
            cell.font = Font(size=14, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0F2B47', end_color='0F2B47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if len(original_columns) > 1:
                worksheet.merge_cells(start_row=current_row, start_column=1, 
                                    end_row=current_row, end_column=len(original_columns))
            
            current_division_type = division_type
            current_row += 3
        
        # Get sex from first row for translation
        first_row = group_data.iloc[0]
        translated_division = translate_division_name(division, first_row['Sex'])
        category_title = f"{translated_division} - {weight_class}kg"
        worksheet.cell(row=current_row, column=1, value=category_title).font = Font(size=12, bold=True)
        current_row += 2
        
        translated_headers = translate_column_headers(original_columns)
        for col_idx, header in enumerate(translated_headers, 1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        for _, row_data in group_data.iterrows():
            place_value = str(row_data['Place']).strip()
            medal_fill = None
            
            if place_value == '1':
                medal_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Gold
            elif place_value == '2':
                medal_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # Silver
            elif place_value == '3':
                medal_fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')  # Bronze
            
            for col_idx, col_name in enumerate(original_columns, 1):
                value = row_data[col_name]
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                if medal_fill is not None:
                    cell.fill = medal_fill
            
            current_row += 1
        
        current_row += 1
    
    auto_fit_columns(worksheet)

def create_club_summary_sheet(worksheet, rankings, header_font, header_fill, header_alignment, data_font, data_alignment, border):
    """Create club rankings summary sheet"""
    
    categories = [
        ('Male_Powerlifting_Ranking.csv', 'Muški Powerlifting'),
        ('Female_Powerlifting_Ranking.csv', 'Ženski Powerlifting'),
        ('Male_Bench_Only_Ranking.csv', 'Muški Potisak s klupe'),
        ('Female_Bench_Only_Ranking.csv', 'Ženski Potisak s klupe')
    ]
    
    current_row = 1
    
    for filename, category_name in categories:
        try:
            df_ranking = rankings.get(filename, pd.read_csv(filename))
            
            if df_ranking.empty:
                continue
            
            # Add category header
            worksheet.cell(row=current_row, column=1, value=f"{category_name} Rang Klubova").font = Font(size=14, bold=True)
            current_row += 2
            
            # Add column headers
            headers = ['Mjesto', 'Klub', 'Bodovi']
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            current_row += 1
            
            # Add data
            for _, row_data in df_ranking.iterrows():
                for col_idx, value in enumerate([int(row_data['Place']), row_data['Club'], row_data['Points']], 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # Medal colors for club rankings
                    if row_data['Place'] == 1:
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Gold
                    elif row_data['Place'] == 2:
                        cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # Silver
                    elif row_data['Place'] == 3:
                        cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')  # Bronze
                
                current_row += 1
            
            current_row += 2  # Space between categories
            
        except Exception as e:
            print(f"Nije moguće učitati {filename}: {e}")
    
    auto_fit_columns(worksheet)

def create_statistics_sheet(worksheet, df, header_font, header_fill, header_alignment, data_font, data_alignment, border):
    """Create statistics summary sheet"""
    
    worksheet.cell(row=1, column=1, value="Statistika Natjecanja - Pitomača Open 2025").font = Font(size=16, bold=True)
    
    current_row = 3
    
    # General statistics
    unique_athletes = df.drop_duplicates(subset=['Name'])
    stats = [
        ("Ukupno Nastupa", len(df)),
        ("Ukupno Natjecatelja", len(unique_athletes)),
        ("Muških Natjecatelja", len(unique_athletes[unique_athletes['Sex'] == 'M'])),
        ("Ženskih Natjecatelja", len(unique_athletes[unique_athletes['Sex'] == 'F'])),
        ("Powerlifting nastupa", len(df[df['Event'] == 'SBD'])),
        ("Potisak s klupe nastupa", len(df[df['Event'] == 'B'])),
        ("Ukupno Klubova", df[df['Club'] != '']['Club'].nunique()),
        ("Prosjek GL Bodova", f"{df['Points'].mean():.2f}"),
        ("Najbolji GL Bodovi", f"{df['Points'].max():.2f}")
    ]
    
    for stat_name, stat_value in stats:
        worksheet.cell(row=current_row, column=1, value=stat_name).font = Font(bold=True)
        stat_cell = worksheet.cell(row=current_row, column=2, value=stat_value)
        stat_cell.alignment = data_alignment
        current_row += 1
    
    current_row += 2
    
    division_order = ['Sub-Junior', 'Junior', 'Open', 'Master I', 'Master II', 'Master III', 'Master IV']
    
    df = df.copy()
    df['DivisionType'] = df['Division'].apply(get_division_type)
    
    def create_top_5_section(title, data, current_row):
        if len(data) == 0:
            return current_row
            
        worksheet.cell(row=current_row, column=1, value=title).font = Font(size=12, bold=True)
        current_row += 2
        
        headers = ['Rang', 'Ime', 'Klub', 'Ukupno (kg)', 'GL Bodovi']
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        top_5 = data.nlargest(5, 'Points')
        for rank, (_, performer) in enumerate(top_5.iterrows(), 1):
            values = [rank, performer['Name'], performer['Club'], performer['TotalKg'], f"{performer['Points']:.2f}"]
            
            for col_idx, value in enumerate(values, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                if rank == 1:
                    cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                elif rank == 2:
                    cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                elif rank == 3:
                    cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
            
            current_row += 1
        
        current_row += 2
        return current_row
    
    # Male Powerlifting
    male_powerlifting = df[(df['Sex'] == 'M') & (df['Event'] == 'SBD')]
    current_row = create_top_5_section("Top 5 Muški Powerlifting", male_powerlifting, current_row)
    
    for division_type in division_order:
        division_data = male_powerlifting[male_powerlifting['DivisionType'] == division_type]
        if len(division_data) > 0:
            translated_division_type = translate_division_type(division_type).title()
            current_row = create_top_5_section(f"Top 5 Muški {translated_division_type} Powerlifting", division_data, current_row)
    
    # Female Powerlifting
    female_powerlifting = df[(df['Sex'] == 'F') & (df['Event'] == 'SBD')]
    current_row = create_top_5_section("Top 5 Ženski Powerlifting", female_powerlifting, current_row)
    
    for division_type in division_order:
        division_data = female_powerlifting[female_powerlifting['DivisionType'] == division_type]
        if len(division_data) > 0:
            translated_division_type = translate_division_type(division_type).title()
            current_row = create_top_5_section(f"Top 5 Ženski {translated_division_type} Powerlifting", division_data, current_row)
    
    # Male Bench Only
    male_bench = df[(df['Sex'] == 'M') & (df['Event'] == 'B')]
    current_row = create_top_5_section("Top 5 Muški Potisak s klupe", male_bench, current_row)
    
    for division_type in division_order:
        division_data = male_bench[male_bench['DivisionType'] == division_type]
        if len(division_data) > 0:
            translated_division_type = translate_division_type(division_type).title()
            current_row = create_top_5_section(f"Top 5 Muški {translated_division_type} Potisak s klupe", division_data, current_row)
    
    # Female Bench Only
    female_bench = df[(df['Sex'] == 'F') & (df['Event'] == 'B')]
    current_row = create_top_5_section("Top 5 Ženski Potisak s klupe", female_bench, current_row)
    
    for division_type in division_order:
        division_data = female_bench[female_bench['DivisionType'] == division_type]
        if len(division_data) > 0:
            translated_division_type = translate_division_type(division_type).title()
            current_row = create_top_5_section(f"Top 5 Ženski {translated_division_type} Potisak s klupe", division_data, current_row)
    
    auto_fit_columns(worksheet)

def create_pitomaca_excel_report():
    """Create a beautifully formatted Excel file for Pitomača competition"""
    
    print("=" * 50)
    print("PITOMAČA OPEN 2025 - Generiranje izvještaja")
    print("=" * 50)
    
    print("\nObrada podataka iz Pitomače...")
    df = process_pitomaca_data()
    
    print(f"\nObrađeno {len(df)} rezultata")
    
    # Generate club rankings
    print("\n" + "-" * 50)
    print("Generiranje poretka klubova...")
    rankings = generate_club_rankings(df)
    
    print("\n" + "-" * 50)
    print("Kreiranje Excel izvještaja...")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 1. Men's Powerlifting Sheet
    print("  Kreiranje 'Muški Powerlifting' stranice...")
    men_sbd = df[(df['Sex'] == 'M') & (df['Event'] == 'SBD')].copy()
    men_sbd = sort_by_categories(men_sbd)
    
    ws_men_sbd = wb.create_sheet("Muški Powerlifting")
    create_formatted_sheet(ws_men_sbd, men_sbd, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 2. Women's Powerlifting Sheet
    print("  Kreiranje 'Ženski Powerlifting' stranice...")
    women_sbd = df[(df['Sex'] == 'F') & (df['Event'] == 'SBD')].copy()
    women_sbd = sort_by_categories(women_sbd)
    
    ws_women_sbd = wb.create_sheet("Ženski Powerlifting")
    create_formatted_sheet(ws_women_sbd, women_sbd, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 3. Men's Bench Only Sheet
    print("  Kreiranje 'Muški Potisak s klupe' stranice...")
    men_bench = df[(df['Sex'] == 'M') & (df['Event'] == 'B')].copy()
    men_bench = sort_by_categories(men_bench)
    
    squat_deadlift_columns = ['Squat1Kg', 'Squat2Kg', 'Squat3Kg', 'Best3SquatKg', 
                             'Deadlift1Kg', 'Deadlift2Kg', 'Deadlift3Kg', 'Best3DeadliftKg']
    men_bench_filtered = men_bench.drop(columns=[col for col in squat_deadlift_columns if col in men_bench.columns])
    
    ws_men_bench = wb.create_sheet("Muški Potisak s klupe")
    create_formatted_sheet(ws_men_bench, men_bench_filtered, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 4. Women's Bench Only Sheet
    print("  Kreiranje 'Ženski Potisak s klupe' stranice...")
    women_bench = df[(df['Sex'] == 'F') & (df['Event'] == 'B')].copy()
    women_bench = sort_by_categories(women_bench)
    
    women_bench_filtered = women_bench.drop(columns=[col for col in squat_deadlift_columns if col in women_bench.columns])
    
    ws_women_bench = wb.create_sheet("Ženski Potisak s klupe")
    create_formatted_sheet(ws_women_bench, women_bench_filtered, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 5. Club Rankings Sheet
    print("  Kreiranje 'Rang Klubova' stranice...")
    ws_clubs = wb.create_sheet("Rang Klubova")
    create_club_summary_sheet(ws_clubs, rankings, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 6. Statistics Sheet
    print("  Kreiranje 'Statistika' stranice...")
    ws_stats = wb.create_sheet("Statistika")
    create_statistics_sheet(ws_stats, df, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    wb.active = wb["Muški Powerlifting"]
    
    filename = "Pitomaca_Open_Rezultati.xlsx"
    wb.save(filename)
    
    print("\n" + "=" * 50)
    print(f"✅ Excel datoteka '{filename}' uspješno kreirana!")
    print(f"📊 Ukupno stranica: {len(wb.sheetnames)}")
    print(f"📈 Ukupno zapisa: {len(df)}")
    print("📋 Stranice: Muški Powerlifting, Ženski Powerlifting, Muški Potisak s klupe, Ženski Potisak s klupe, Rang Klubova, Statistika")
    print("=" * 50)
    
    return filename

if __name__ == "__main__":
    try:
        create_pitomaca_excel_report()
    except Exception as e:
        print(f"Greška pri kreiranju Excel datoteke: {e}")
        import traceback
        traceback.print_exc()
