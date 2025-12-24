import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import math

def calculate_ipf_gl_points(bodyweight, total, sex):
    """Calculate IPF GL Points using official IPF formula for Bench Only"""
    if pd.isna(bodyweight) or pd.isna(total) or total == 0:
        return 0
    
    # Official IPF GL Coefficients for Classic Bench Press
    if sex == 'M':  # Men's Classic Bench Press
        A = 320.98041
        B = 281.40258
        C = 0.01008
    else:  # Women's Classic Bench Press
        A = 142.40398
        B = 442.52671
        C = 0.04724
    
    try:
        # IPF GL Points formula: 100/(A-B*exp(-C*bodyweight)) * total
        ipf_gl_points = 100 / (A - B * math.exp(-C * bodyweight)) * total
        return round(ipf_gl_points, 2)
    except:
        return 0

def load_club_mapping():
    """Load club and birth year mapping from klubovi.csv"""
    club_mapping = {}
    birth_year_mapping = {}
    
    try:
        df = pd.read_csv('zagreb/klubovi.csv', header=None)
        
        for _, row in df.iterrows():
            try:
                ime = row[2]
                prezime = row[3]
                godiste = row[4]
                klub = row[5]
                
                if pd.notna(ime) and pd.notna(prezime) and str(ime).strip() != '' and str(prezime).strip() != '':
                    # Skip header rows
                    if str(ime).upper() == 'IME' or str(prezime).upper() == 'PREZIME':
                        continue
                    
                    full_name = f"{str(ime).strip()} {str(prezime).strip()}"
                    
                    if pd.notna(klub) and str(klub).strip() != '':
                        club_mapping[full_name] = str(klub).strip()
                    
                    if pd.notna(godiste):
                        try:
                            year = int(float(godiste))
                            birth_year_mapping[full_name] = year
                        except:
                            pass
            except:
                continue
        
        print(f"Učitano {len(club_mapping)} mapiranja klubova")
        print(f"Učitano {len(birth_year_mapping)} mapiranja godina rođenja")
        
    except Exception as e:
        print(f"Greška pri učitavanju klubovi.csv: {e}")
    
    return club_mapping, birth_year_mapping

def is_equipped(division_name):
    """Check if division is equipped (EQ)"""
    return '-EQ' in division_name or 'EQ' in division_name.upper().split('-')

def is_osi(division_name):
    """Check if division is OSI (Paralympic)"""
    return '-OSI' in division_name or 'OSI' in division_name.upper().split('-')

def get_division_type(division_name):
    """Extract division type from full division name"""
    # Remove EQ suffix for matching
    clean_name = division_name.replace('-EQ', '').replace(' EQ', '')
    
    if 'Master 3' in clean_name or 'Master III' in clean_name:
        return 'Master III'
    elif 'Master 2' in clean_name or 'Master II' in clean_name:
        return 'Master II'
    elif 'Master 1' in clean_name or 'Master I' in clean_name:
        return 'Master I'
    elif 'Kadet' in clean_name or 'Sub-Junior' in clean_name:
        return 'Sub-Junior'
    elif 'Junior' in clean_name and 'Sub' not in clean_name:
        return 'Junior'
    elif 'Open' in clean_name:
        return 'Open'
    else:
        return 'Open'  # Default

def translate_column_headers(columns):
    """Translate English column headers to Croatian"""
    translation_map = {
        'Place': 'Plasman',
        'Name': 'Ime i prezime',
        'Club': 'Klub',
        'Sex': 'Spol',
        'BirthYear': 'Godina rođenja',
        'Division': 'Kategorija',
        'BodyweightKg': 'Tjelesna masa (kg)',
        'WeightClassKg': 'Težinska kategorija (kg)',
        'Bench1Kg': 'Potisak s klupe 1 (kg)',
        'Bench2Kg': 'Potisak s klupe 2 (kg)',
        'Bench3Kg': 'Potisak s klupe 3 (kg)',
        'Best3BenchKg': 'Najbolji potisak s klupe (kg)',
        'TotalKg': 'Ukupno (kg)',
        'Points': 'GL Bodovi',
        'Event': 'Disciplina'
    }
    
    return [translation_map.get(col, col) for col in columns]

def translate_division_name(division_name, sex, is_eq=False):
    """Translate English division names to Croatian with gender and EQ suffix"""
    gender_prefix = "Muški " if sex == 'M' else "Ženski "
    eq_suffix = " EQ" if is_eq else ""
    
    # Remove EQ from division name for matching
    clean_name = division_name.replace('-EQ', '').replace(' EQ', '')
    
    # Translate division types
    if 'Master 3' in clean_name or 'Master III' in clean_name:
        return gender_prefix + "Veterani 3" + eq_suffix
    elif 'Master 2' in clean_name or 'Master II' in clean_name:
        return gender_prefix + "Veterani 2" + eq_suffix
    elif 'Master 1' in clean_name or 'Master I' in clean_name:
        return gender_prefix + "Veterani 1" + eq_suffix
    elif 'Kadet' in clean_name or 'Sub-Junior' in clean_name:
        return gender_prefix + "Kadeti" + eq_suffix
    elif 'Junior' in clean_name and 'Sub' not in clean_name:
        return gender_prefix + "Juniori" + eq_suffix
    elif 'Open' in clean_name:
        return gender_prefix + "Seniori" + eq_suffix
    else:
        return gender_prefix + "Seniori" + eq_suffix  # Default

def translate_division_type(division_type, is_eq=False):
    """Translate division type for headers"""
    translations = {
        'Sub-Junior': 'KADETI',
        'Junior': 'JUNIORI',
        'Open': 'SENIORI',
        'Master I': 'VETERANI 1',
        'Master II': 'VETERANI 2', 
        'Master III': 'VETERANI 3',
        'Master IV': 'VETERANI 4'
    }
    base = translations.get(division_type, division_type.upper())
    if is_eq:
        return base + " EQ"
    return base

def auto_fit_columns(worksheet):
    """Auto-fit all column widths in the worksheet"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        
        for cell in column:
            try:
                if hasattr(cell, 'column_letter') and cell.column_letter:
                    if column_letter is None:
                        column_letter = cell.column_letter
                    
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            except:
                pass
        
        if column_letter is not None:
            adjusted_width = min(max_length + 3, 30)
            if adjusted_width < 8:
                adjusted_width = 8
            worksheet.column_dimensions[column_letter].width = adjusted_width

def process_zagreb_data():
    """Process Zagreb CSV data into standardized format - BENCH ONLY"""
    
    # Load club and birth year mappings
    club_mapping, birth_year_mapping = load_club_mapping()
    
    # Read the CSV - skip the first 5 rows (header info)
    df = pd.read_csv('zagreb/rez.csv', skiprows=5)
    
    print(f"Ukupno redaka u CSV: {len(df)}")
    
    # Filter for Bench Only events only
    df_bench = df[df['Event'] == 'B'].copy()
    print(f"Bench Only rezultata: {len(df_bench)}")
    
    # Filter out non-valid results (NS = No Show, DQ = Disqualified)
    df_valid = df_bench[~df_bench['Place'].isin(['NS', 'DQ', ''])]
    df_valid = df_valid.dropna(subset=['Place'])
    
    # Convert Place to numeric where possible
    df_valid = df_valid.copy()
    df_valid['PlaceNum'] = pd.to_numeric(df_valid['Place'], errors='coerce')
    df_valid = df_valid[df_valid['PlaceNum'].notna()]
    
    print(f"Važeći rezultati: {len(df_valid)}")
    
    # Filter out OSI (Paralympic) competitors - they don't count in rankings
    valid_before_osi = len(df_valid)
    df_valid = df_valid[~df_valid['Division'].apply(is_osi)]
    osi_excluded = valid_before_osi - len(df_valid)
    if osi_excluded > 0:
        print(f"Isključeno OSI natjecatelja: {osi_excluded}")
    
    print(f"Rezultati za poredak: {len(df_valid)}")
    
    # Create standardized output
    output_data = []
    
    for _, row in df_valid.iterrows():
        sex = row['Sex']
        division = row['Division']
        name = row['Name']
        
        # Check if this is equipped
        eq = is_equipped(division)
        
        # Create Division name
        gender = "Men's" if sex == 'M' else "Women's"
        eq_tag = " EQ" if eq else ""
        # Clean division name (remove -EQ suffix)
        clean_division = division.replace('-EQ', '').replace('-OSI', '')
        full_division = f"{gender} Raw {clean_division}{eq_tag} Bench Only"
        
        # Get best bench for total
        best_bench = row['Best3BenchKg'] if pd.notna(row['Best3BenchKg']) else 0
        
        # Calculate GL points (use existing Points if available, otherwise calculate)
        gl_points = row['Points'] if pd.notna(row['Points']) and row['Points'] > 0 else calculate_ipf_gl_points(
            bodyweight=pd.to_numeric(row['BodyweightKg'], errors='coerce'),
            total=best_bench,
            sex=sex
        )
        
        # Get club from Team column in CSV or mapping
        club = row['Team'] if pd.notna(row['Team']) and str(row['Team']).strip() != '' else club_mapping.get(name, '')
        
        # Get birth year from mapping
        birth_year = birth_year_mapping.get(name, '')
        
        output_data.append({
            'Place': int(row['PlaceNum']),
            'Name': name,
            'Club': club,
            'Sex': sex,
            'BirthYear': birth_year,
            'Division': full_division,
            'BodyweightKg': row['BodyweightKg'],
            'WeightClassKg': row['WeightClassKg'],
            'Bench1Kg': row['Bench1Kg'] if pd.notna(row['Bench1Kg']) else '',
            'Bench2Kg': row['Bench2Kg'] if pd.notna(row['Bench2Kg']) else '',
            'Bench3Kg': row['Bench3Kg'] if pd.notna(row['Bench3Kg']) else '',
            'Best3BenchKg': row['Best3BenchKg'] if pd.notna(row['Best3BenchKg']) else '',
            'TotalKg': best_bench,
            'Points': gl_points,
            'IsEquipped': eq
        })
    
    result_df = pd.DataFrame(output_data)
    
    # Check for missing club or birth year data - raise error if found
    missing_data = []
    for _, row in result_df.iterrows():
        name = row['Name']
        missing = []
        if row['Club'] == '':
            missing.append('klub')
        if row['BirthYear'] == '':
            missing.append('godina rođenja')
        if missing:
            missing_data.append(f"  - {name}: nedostaje {', '.join(missing)}")
    
    if missing_data:
        error_msg = f"\n❌ GREŠKA: Nedostaju podaci za {len(missing_data)} natjecatelja:\n"
        error_msg += "\n".join(missing_data)
        error_msg += "\n\nMolimo ažurirajte zagreb/klubovi.csv datoteku s podacima koji nedostaju."
        raise ValueError(error_msg)
    
    print(f"✓ Pronađeni svi podaci (klub i godina rođenja) za {len(result_df)} natjecatelja")
    
    # Report equipped count
    eq_count = result_df['IsEquipped'].sum()
    raw_count = len(result_df) - eq_count
    print(f"  - RAW natjecatelja: {raw_count}")
    print(f"  - EQ natjecatelja: {eq_count}")
    
    return result_df

def sort_by_categories(df):
    """Sort dataframe by division order, then weight class, then place"""
    
    division_order = {
        'Sub-Junior': 1,
        'Junior': 2, 
        'Open': 3,
        'Master I': 4,
        'Master II': 5,
        'Master III': 6,
        'Master IV': 7
    }
    
    df = df.copy()
    df['DivisionOrder'] = df['Division'].apply(lambda x: division_order.get(get_division_type(x), 3))
    
    def weight_sort_key(weight_class_str):
        weight_str = str(weight_class_str)
        try:
            if '+' in weight_str:
                base_weight = float(weight_str.replace('+', ''))
                return base_weight + 0.5
            else:
                return float(weight_str)
        except ValueError:
            return 9999.0
    
    df['WeightClassKg_num'] = df['WeightClassKg'].apply(weight_sort_key)
    df['Place_num'] = pd.to_numeric(df['Place'], errors='coerce')
    
    df_sorted = df.sort_values(['DivisionOrder', 'WeightClassKg_num', 'Place_num'])
    df_sorted = df_sorted.drop(['DivisionOrder', 'WeightClassKg_num', 'Place_num'], axis=1)
    
    return df_sorted

def generate_club_rankings(df, is_eq=False):
    """Generate club ranking CSV files based on top 5 lifters per club"""
    
    eq_suffix = "_EQ" if is_eq else ""
    eq_label = " EQ" if is_eq else ""
    
    categories = [
        (df[df['Sex'] == 'M'], f'Male_Bench_Only{eq_suffix}_Ranking.csv', f'Muški Potisak s klupe{eq_label}'),
        (df[df['Sex'] == 'F'], f'Female_Bench_Only{eq_suffix}_Ranking.csv', f'Ženski Potisak s klupe{eq_label}')
    ]
    
    rankings = {}
    
    for data, filename, category_name in categories:
        print(f"\nGeneriranje rangiranja za {category_name}...")
        
        # Filter out empty clubs
        data_with_clubs = data[data['Club'] != ''].copy()
        
        if data_with_clubs.empty:
            print(f"  Nema podataka o klubovima za {category_name}")
            empty_df = pd.DataFrame(columns=['Place', 'Club', 'Points'])
            empty_df.to_csv(f'zagreb/{filename}', index=False)
            rankings[filename] = empty_df
            continue
        
        # For each club, take only the top 5 lifters by Points
        def get_top_5_points(group):
            top_5 = group.nlargest(5, 'Points')
            return top_5['Points'].sum()
        
        club_totals = data_with_clubs.groupby('Club').apply(get_top_5_points, include_groups=False).reset_index()
        club_totals.columns = ['Club', 'Points']
        
        # Sort by total points (descending) and add ranking
        club_totals = club_totals.sort_values('Points', ascending=False).reset_index(drop=True)
        club_totals['Place'] = range(1, len(club_totals) + 1)
        
        # Reorder columns
        club_rankings = club_totals[['Place', 'Club', 'Points']].copy()
        club_rankings['Points'] = club_rankings['Points'].round(2)
        
        # Save the ranking
        club_rankings.to_csv(f'zagreb/{filename}', index=False)
        rankings[filename] = club_rankings
        
        print(f"  Generirano {len(club_rankings)} klubova (top 5 liftera po klubu)")
        
        # Show top 3 clubs
        if len(club_rankings) > 0:
            print("  Top 3 klubova:")
            for _, row in club_rankings.head(3).iterrows():
                print(f"    {int(row['Place'])}. {row['Club']} - {row['Points']:.2f} bodova")
    
    return rankings

def create_formatted_sheet(worksheet, data, header_font, header_fill, header_alignment, data_font, data_alignment, border, is_eq=False):
    """Create a formatted sheet with separate tables for each division/weight class combination"""
    
    if len(data) == 0:
        return
    
    current_row = 1
    division_order = ['Sub-Junior', 'Junior', 'Open', 'Master I', 'Master II', 'Master III', 'Master IV']
    
    data_copy = data.copy()
    data_copy['DivisionType'] = data_copy['Division'].apply(get_division_type)
    
    def weight_sort_key(weight_class_str):
        weight_str = str(weight_class_str)
        try:
            if '+' in weight_str:
                base_weight = float(weight_str.replace('+', ''))
                return base_weight + 0.5
            else:
                return float(weight_str)
        except ValueError:
            return 9999.0
    
    data_copy['WeightSortKey'] = data_copy['WeightClassKg'].apply(weight_sort_key)
    data_copy['PlaceNumeric'] = pd.to_numeric(data_copy['Place'], errors='coerce')
    
    division_order_map = {div: i for i, div in enumerate(division_order)}
    data_copy['DivisionOrder'] = data_copy['DivisionType'].map(division_order_map).fillna(999)
    data_sorted = data_copy.sort_values(['DivisionOrder', 'WeightSortKey', 'PlaceNumeric'])
    
    # Get original column names (excluding helper columns, Division, WeightClassKg, Sex, IsEquipped)
    original_columns = [col for col in data.columns if col not in ['Division', 'WeightClassKg', 'Sex', 'IsEquipped']]
    
    # Get unique division/weight class combinations in sorted order
    unique_combinations = []
    seen = set()
    for _, row in data_sorted.iterrows():
        combo = (row['Division'], row['WeightClassKg'])
        if combo not in seen:
            unique_combinations.append(combo)
            seen.add(combo)
    
    current_division_type = None
    
    for division, weight_class in unique_combinations:
        group_data = data_sorted[(data_sorted['Division'] == division) & 
                                (data_sorted['WeightClassKg'] == weight_class)]
        
        division_type = get_division_type(division)
        
        if current_division_type != division_type:
            if current_division_type is not None:
                current_row += 2
            
            translated_division_type = translate_division_type(division_type, is_eq)
            division_header = f"═══ {translated_division_type} KATEGORIJA ═══"
            cell = worksheet.cell(row=current_row, column=1, value=division_header)
            cell.font = Font(size=14, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0F2B47', end_color='0F2B47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if len(original_columns) > 1:
                worksheet.merge_cells(start_row=current_row, start_column=1, 
                                    end_row=current_row, end_column=len(original_columns))
            
            current_division_type = division_type
            current_row += 3
        
        # Get sex from first row for translation
        first_row = group_data.iloc[0]
        translated_division = translate_division_name(division, first_row['Sex'], is_eq)
        category_title = f"{translated_division} - {weight_class}kg"
        worksheet.cell(row=current_row, column=1, value=category_title).font = Font(size=12, bold=True)
        current_row += 2
        
        translated_headers = translate_column_headers(original_columns)
        for col_idx, header in enumerate(translated_headers, 1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        for _, row_data in group_data.iterrows():
            place_value = str(row_data['Place']).strip()
            medal_fill = None
            
            if place_value == '1':
                medal_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Gold
            elif place_value == '2':
                medal_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # Silver
            elif place_value == '3':
                medal_fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')  # Bronze
            
            for col_idx, col_name in enumerate(original_columns, 1):
                value = row_data[col_name]
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                if medal_fill is not None:
                    cell.fill = medal_fill
            
            current_row += 1
        
        current_row += 1
    
    auto_fit_columns(worksheet)

def create_club_summary_sheet(worksheet, rankings_raw, rankings_eq, header_font, header_fill, header_alignment, data_font, data_alignment, border):
    """Create club rankings summary sheet - BENCH ONLY (RAW + EQ)"""
    
    current_row = 1
    
    # RAW Rankings
    worksheet.cell(row=current_row, column=1, value="═══ RAW POREDAK KLUBOVA ═══").font = Font(size=16, bold=True)
    current_row += 2
    
    raw_categories = [
        ('Male_Bench_Only_Ranking.csv', 'Muški Potisak s klupe'),
        ('Female_Bench_Only_Ranking.csv', 'Ženski Potisak s klupe')
    ]
    
    for filename, category_name in raw_categories:
        try:
            df_ranking = rankings_raw.get(filename, pd.DataFrame())
            
            if df_ranking.empty:
                continue
            
            # Add category header
            worksheet.cell(row=current_row, column=1, value=f"{category_name} Rang Klubova").font = Font(size=14, bold=True)
            current_row += 2
            
            # Add column headers
            headers = ['Mjesto', 'Klub', 'Bodovi']
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            current_row += 1
            
            # Add data
            for _, row_data in df_ranking.iterrows():
                for col_idx, value in enumerate([int(row_data['Place']), row_data['Club'], row_data['Points']], 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # Medal colors for club rankings
                    if row_data['Place'] == 1:
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                    elif row_data['Place'] == 2:
                        cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                    elif row_data['Place'] == 3:
                        cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                
                current_row += 1
            
            current_row += 2
            
        except Exception as e:
            print(f"Nije moguće učitati {filename}: {e}")
    
    # EQ Rankings (if any)
    if rankings_eq:
        current_row += 2
        worksheet.cell(row=current_row, column=1, value="═══ EQUIPPED (EQ) POREDAK KLUBOVA ═══").font = Font(size=16, bold=True)
        current_row += 2
        
        eq_categories = [
            ('Male_Bench_Only_EQ_Ranking.csv', 'Muški Potisak s klupe EQ'),
            ('Female_Bench_Only_EQ_Ranking.csv', 'Ženski Potisak s klupe EQ')
        ]
        
        for filename, category_name in eq_categories:
            try:
                df_ranking = rankings_eq.get(filename, pd.DataFrame())
                
                if df_ranking.empty:
                    continue
                
                # Add category header
                worksheet.cell(row=current_row, column=1, value=f"{category_name} Rang Klubova").font = Font(size=14, bold=True)
                current_row += 2
                
                # Add column headers
                headers = ['Mjesto', 'Klub', 'Bodovi']
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                current_row += 1
                
                # Add data
                for _, row_data in df_ranking.iterrows():
                    for col_idx, value in enumerate([int(row_data['Place']), row_data['Club'], row_data['Points']], 1):
                        cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = border
                        
                        # Medal colors
                        if row_data['Place'] == 1:
                            cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        elif row_data['Place'] == 2:
                            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                        elif row_data['Place'] == 3:
                            cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                    
                    current_row += 1
                
                current_row += 2
                
            except Exception as e:
                print(f"Nije moguće učitati {filename}: {e}")
    
    auto_fit_columns(worksheet)

def create_statistics_sheet(worksheet, df_raw, df_eq, header_font, header_fill, header_alignment, data_font, data_alignment, border):
    """Create statistics summary sheet - BENCH ONLY (RAW + EQ)"""
    
    worksheet.cell(row=1, column=1, value="Statistika Natjecanja - 14. Državno prvenstvo Bench Press Zagreb 2025").font = Font(size=16, bold=True)
    
    current_row = 3
    
    # Combined stats
    df_all = pd.concat([df_raw, df_eq]) if len(df_eq) > 0 else df_raw
    unique_athletes = df_all.drop_duplicates(subset=['Name'])
    
    stats = [
        ("Ukupno Nastupa (Bench Only)", len(df_all)),
        ("  - RAW nastupa", len(df_raw)),
        ("  - EQ nastupa", len(df_eq)),
        ("Ukupno Natjecatelja", len(unique_athletes)),
        ("Muških Natjecatelja", len(unique_athletes[unique_athletes['Sex'] == 'M'])),
        ("Ženskih Natjecatelja", len(unique_athletes[unique_athletes['Sex'] == 'F'])),
        ("Ukupno Klubova", df_all[df_all['Club'] != '']['Club'].nunique()),
    ]
    
    for stat_name, stat_value in stats:
        worksheet.cell(row=current_row, column=1, value=stat_name).font = Font(bold=True)
        stat_cell = worksheet.cell(row=current_row, column=2, value=stat_value)
        stat_cell.alignment = data_alignment
        current_row += 1
    
    current_row += 2
    
    division_order = ['Sub-Junior', 'Junior', 'Open', 'Master I', 'Master II', 'Master III', 'Master IV']
    
    def create_top_5_section(title, data, current_row):
        if len(data) == 0:
            return current_row
        
        data = data.copy()
        data['DivisionType'] = data['Division'].apply(get_division_type)
            
        worksheet.cell(row=current_row, column=1, value=title).font = Font(size=12, bold=True)
        current_row += 2
        
        headers = ['Rang', 'Ime', 'Klub', 'Potisak (kg)', 'GL Bodovi']
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        top_5 = data.nlargest(5, 'Points')
        for rank, (_, performer) in enumerate(top_5.iterrows(), 1):
            values = [rank, performer['Name'], performer['Club'], performer['TotalKg'], f"{performer['Points']:.2f}"]
            
            for col_idx, value in enumerate(values, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                if rank == 1:
                    cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                elif rank == 2:
                    cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                elif rank == 3:
                    cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
            
            current_row += 1
        
        current_row += 2
        return current_row
    
    # RAW Statistics
    worksheet.cell(row=current_row, column=1, value="═══ RAW STATISTIKA ═══").font = Font(size=14, bold=True)
    current_row += 2
    
    if len(df_raw) > 0:
        worksheet.cell(row=current_row, column=1, value=f"Prosjek GL Bodova (RAW): {df_raw['Points'].mean():.2f}").font = Font(bold=True)
        current_row += 1
        worksheet.cell(row=current_row, column=1, value=f"Najbolji GL Bodovi (RAW): {df_raw['Points'].max():.2f}").font = Font(bold=True)
        current_row += 2
        
        df_raw_copy = df_raw.copy()
        df_raw_copy['DivisionType'] = df_raw_copy['Division'].apply(get_division_type)
        
        # Male RAW
        male_bench = df_raw_copy[df_raw_copy['Sex'] == 'M']
        current_row = create_top_5_section("Top 5 Muški Potisak s klupe (RAW)", male_bench, current_row)
        
        for division_type in division_order:
            division_data = male_bench[male_bench['DivisionType'] == division_type]
            if len(division_data) > 0:
                translated_division_type = translate_division_type(division_type).title()
                current_row = create_top_5_section(f"Top 5 Muški {translated_division_type} Potisak s klupe (RAW)", division_data, current_row)
        
        # Female RAW
        female_bench = df_raw_copy[df_raw_copy['Sex'] == 'F']
        current_row = create_top_5_section("Top 5 Ženski Potisak s klupe (RAW)", female_bench, current_row)
        
        for division_type in division_order:
            division_data = female_bench[female_bench['DivisionType'] == division_type]
            if len(division_data) > 0:
                translated_division_type = translate_division_type(division_type).title()
                current_row = create_top_5_section(f"Top 5 Ženski {translated_division_type} Potisak s klupe (RAW)", division_data, current_row)
    
    # EQ Statistics
    if len(df_eq) > 0:
        current_row += 2
        worksheet.cell(row=current_row, column=1, value="═══ EQUIPPED (EQ) STATISTIKA ═══").font = Font(size=14, bold=True)
        current_row += 2
        
        worksheet.cell(row=current_row, column=1, value=f"Prosjek GL Bodova (EQ): {df_eq['Points'].mean():.2f}").font = Font(bold=True)
        current_row += 1
        worksheet.cell(row=current_row, column=1, value=f"Najbolji GL Bodovi (EQ): {df_eq['Points'].max():.2f}").font = Font(bold=True)
        current_row += 2
        
        df_eq_copy = df_eq.copy()
        df_eq_copy['DivisionType'] = df_eq_copy['Division'].apply(get_division_type)
        
        # Male EQ
        male_bench_eq = df_eq_copy[df_eq_copy['Sex'] == 'M']
        current_row = create_top_5_section("Top 5 Muški Potisak s klupe (EQ)", male_bench_eq, current_row)
        
        for division_type in division_order:
            division_data = male_bench_eq[male_bench_eq['DivisionType'] == division_type]
            if len(division_data) > 0:
                translated_division_type = translate_division_type(division_type, is_eq=True).title()
                current_row = create_top_5_section(f"Top 5 Muški {translated_division_type} Potisak s klupe", division_data, current_row)
        
        # Female EQ
        female_bench_eq = df_eq_copy[df_eq_copy['Sex'] == 'F']
        current_row = create_top_5_section("Top 5 Ženski Potisak s klupe (EQ)", female_bench_eq, current_row)
        
        for division_type in division_order:
            division_data = female_bench_eq[female_bench_eq['DivisionType'] == division_type]
            if len(division_data) > 0:
                translated_division_type = translate_division_type(division_type, is_eq=True).title()
                current_row = create_top_5_section(f"Top 5 Ženski {translated_division_type} Potisak s klupe", division_data, current_row)
    
    auto_fit_columns(worksheet)

def create_zagreb_excel_report():
    """Create a beautifully formatted Excel file for Zagreb Bench Only competition"""
    
    print("=" * 60)
    print("14. DRŽAVNO PRVENSTVO BENCH PRESS ZAGREB 2025")
    print("Generiranje izvještaja - BENCH ONLY (RAW + EQ)")
    print("=" * 60)
    
    print("\nObrada podataka iz Zagreba...")
    df = process_zagreb_data()
    
    # Separate RAW and EQ data
    df_raw = df[df['IsEquipped'] == False].copy()
    df_eq = df[df['IsEquipped'] == True].copy()
    
    print(f"\nObrađeno {len(df)} bench only rezultata")
    print(f"  - RAW: {len(df_raw)}")
    print(f"  - EQ: {len(df_eq)}")
    
    # Generate club rankings - RAW
    print("\n" + "-" * 60)
    print("Generiranje poretka klubova (RAW)...")
    rankings_raw = generate_club_rankings(df_raw, is_eq=False)
    
    # Generate club rankings - EQ (if any)
    rankings_eq = {}
    if len(df_eq) > 0:
        print("\nGeneriranje poretka klubova (EQ)...")
        rankings_eq = generate_club_rankings(df_eq, is_eq=True)
    
    print("\n" + "-" * 60)
    print("Kreiranje Excel izvještaja...")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 1. Men's Bench Only RAW Sheet
    print("  Kreiranje 'Muški Potisak s klupe' stranice...")
    men_bench_raw = df_raw[df_raw['Sex'] == 'M'].copy()
    men_bench_raw = sort_by_categories(men_bench_raw)
    
    ws_men_bench = wb.create_sheet("Muški Potisak s klupe")
    create_formatted_sheet(ws_men_bench, men_bench_raw, header_font, header_fill, header_alignment, data_font, data_alignment, border, is_eq=False)
    
    # 2. Women's Bench Only RAW Sheet
    print("  Kreiranje 'Ženski Potisak s klupe' stranice...")
    women_bench_raw = df_raw[df_raw['Sex'] == 'F'].copy()
    women_bench_raw = sort_by_categories(women_bench_raw)
    
    ws_women_bench = wb.create_sheet("Ženski Potisak s klupe")
    create_formatted_sheet(ws_women_bench, women_bench_raw, header_font, header_fill, header_alignment, data_font, data_alignment, border, is_eq=False)
    
    # 3. Men's Bench Only EQ Sheet (if any)
    if len(df_eq[df_eq['Sex'] == 'M']) > 0:
        print("  Kreiranje 'Muški Potisak s klupe EQ' stranice...")
        men_bench_eq = df_eq[df_eq['Sex'] == 'M'].copy()
        men_bench_eq = sort_by_categories(men_bench_eq)
        
        ws_men_bench_eq = wb.create_sheet("Muški Potisak s klupe EQ")
        create_formatted_sheet(ws_men_bench_eq, men_bench_eq, header_font, header_fill, header_alignment, data_font, data_alignment, border, is_eq=True)
    
    # 4. Women's Bench Only EQ Sheet (if any)
    if len(df_eq[df_eq['Sex'] == 'F']) > 0:
        print("  Kreiranje 'Ženski Potisak s klupe EQ' stranice...")
        women_bench_eq = df_eq[df_eq['Sex'] == 'F'].copy()
        women_bench_eq = sort_by_categories(women_bench_eq)
        
        ws_women_bench_eq = wb.create_sheet("Ženski Potisak s klupe EQ")
        create_formatted_sheet(ws_women_bench_eq, women_bench_eq, header_font, header_fill, header_alignment, data_font, data_alignment, border, is_eq=True)
    
    # 5. Club Rankings Sheet
    print("  Kreiranje 'Rang Klubova' stranice...")
    ws_clubs = wb.create_sheet("Rang Klubova")
    create_club_summary_sheet(ws_clubs, rankings_raw, rankings_eq, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 6. Statistics Sheet
    print("  Kreiranje 'Statistika' stranice...")
    ws_stats = wb.create_sheet("Statistika")
    create_statistics_sheet(ws_stats, df_raw, df_eq, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    wb.active = wb["Muški Potisak s klupe"]
    
    filename = "zagreb/Zagreb_Bench_Rezultati.xlsx"
    wb.save(filename)
    
    sheet_list = [ws.title for ws in wb.worksheets]
    
    print("\n" + "=" * 60)
    print(f"✅ Excel datoteka '{filename}' uspješno kreirana!")
    print(f"📊 Ukupno stranica: {len(wb.sheetnames)}")
    print(f"📈 Ukupno bench only zapisa: {len(df)} (RAW: {len(df_raw)}, EQ: {len(df_eq)})")
    print(f"📋 Stranice: {', '.join(sheet_list)}")
    print("=" * 60)
    
    return filename

if __name__ == "__main__":
    try:
        create_zagreb_excel_report()
    except Exception as e:
        print(f"Greška pri kreiranju Excel datoteke: {e}")
        import traceback
        traceback.print_exc()
