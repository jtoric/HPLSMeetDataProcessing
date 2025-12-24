"""
Excel Report Generator for Powerlifting Competition Results

This module creates formatted Excel reports from processed powerlifting data,
including individual results, club rankings, and statistics.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def get_division_type(division_name):
    """Extract division type from full division name"""
    # Normalize and detect division types, supporting plural and numeric Masters labels
    # Handle NaN/float values
    if pd.isna(division_name):
        return 'Open'
    text = str(division_name).strip() or ''
    
    # Normalize text for comparison
    text_lower = text.lower()
    
    # Check for Master divisions (various formats)
    if 'master iv' in text_lower or 'masters 4' in text_lower or text == 'Master 4':
        return 'Master IV'
    elif 'master iii' in text_lower or 'masters 3' in text_lower or text == 'Master 3':
        return 'Master III'
    elif 'master ii' in text_lower or 'masters 2' in text_lower or text == 'Master 2':
        return 'Master II'
    elif 'master i' in text_lower or 'masters 1' in text_lower or text == 'Master 1':
        return 'Master I'
    # Check for Sub-Junior/Kadet (various formats)
    elif 'sub-junior' in text_lower or 'sub-juniors' in text_lower or text_lower == 'kadet' or text == 'Kadet':
        return 'Sub-Junior'
    # Check for Junior (but not Sub-Junior)
    elif (('junior' in text_lower) or ('juniors' in text_lower)) and ('sub-junior' not in text_lower and 'sub-juniors' not in text_lower and text_lower != 'kadet'):
        return 'Junior'
    # Check for Open
    elif 'open' in text_lower or 'guest' in text_lower:
        return 'Open'
    else:
        return 'Open'

def translate_column_headers(columns):
    """Translate English column headers to Croatian"""
    translation_map = {
        'Place': 'Plasman',
        'Name': 'Ime i prezime',
        'Club': 'Klub',
        'Sex': 'Spol',
        'BirthYear': 'Godina rođenja',
        'Division': 'Kategorija',
        'BodyweightKg': 'Tjelesna masa (kg)',
        'WeightClassKg': 'Težinska kategorija (kg)',
        'Squat1Kg': 'Čučanj 1 (kg)',
        'Squat2Kg': 'Čučanj 2 (kg)',
        'Squat3Kg': 'Čučanj 3 (kg)',
        'Best3SquatKg': 'Najbolji čučanj (kg)',
        'Bench1Kg': 'Potisak s klupe 1 (kg)',
        'Bench2Kg': 'Potisak s klupe 2 (kg)',
        'Bench3Kg': 'Potisak s klupe 3 (kg)',
        'Best3BenchKg': 'Najbolji potisak s klupe (kg)',
        'Deadlift1Kg': 'Mrtvo dizanje 1 (kg)',
        'Deadlift2Kg': 'Mrtvo dizanje 2 (kg)',
        'Deadlift3Kg': 'Mrtvo dizanje 3 (kg)',
        'Best3DeadliftKg': 'Najbolje mrtvo dizanje (kg)',
        'TotalKg': 'Ukupno (kg)',
        'Points': 'GL Bodovi',
        'Event': 'Disciplina'
    }
    
    return [translation_map.get(col, col) for col in columns]

def translate_division_name(division_name):
    """Translate English division names to Croatian"""
    # Handle NaN/float values
    if pd.isna(division_name):
        return 'Open'
    division_name = str(division_name)
    # Handle gender prefixes
    gender_prefix = ""
    if division_name.startswith("Men's"):
        gender_prefix = "Muški "
        division_name = division_name.replace("Men's ", "")
    elif division_name.startswith("Women's"):
        gender_prefix = "Ženski "
        division_name = division_name.replace("Women's ", "")
    
    # Remove "Raw" as it's implied in Croatian powerlifting
    division_name = division_name.replace("Raw ", "")
    
    # Handle bench only first
    bench_only = ""
    if "Bench Only" in division_name:
        bench_only = " Potisak s klupe"
        division_name = division_name.replace(" Bench Only", "")
    
    # Normalize Masters numeric to roman for translation
    division_name = division_name.replace('Masters 4', 'Master IV').replace('Masters 3', 'Master III').replace('Masters 2', 'Master II').replace('Masters 1', 'Master I')
    
    # Translate division types (order matters)
    if 'Master IV' in division_name:
        division_name = division_name.replace('Master IV', 'Veterani 4')
    elif 'Master III' in division_name:
        division_name = division_name.replace('Master III', 'Veterani 3')
    elif 'Master II' in division_name:
        division_name = division_name.replace('Master II', 'Veterani 2')
    elif 'Master I' in division_name:
        division_name = division_name.replace('Master I', 'Veterani 1')
    elif 'Sub-Junior' in division_name or 'Sub-Juniors' in division_name:
        division_name = division_name.replace('Sub-Juniors', 'Kadeti').replace('Sub-Junior', 'Kadeti')
    elif 'Junior' in division_name or 'Juniors' in division_name:
        division_name = division_name.replace('Juniors', 'Juniori')
    elif 'Open' in division_name:
        division_name = division_name.replace('Open', 'Seniori')
    elif 'Guest' in division_name:
        division_name = division_name.replace('Guest', 'Gost')
    
    return gender_prefix + division_name + bench_only

def translate_division_type(division_type):
    """Translate division type for headers"""
    translations = {
        'Sub-Junior': 'KADETI',
        'Junior': 'JUNIORI',
        'Open': 'SENIORI',
        'Master I': 'VETERANI 1',
        'Master II': 'VETERANI 2', 
        'Master III': 'VETERANI 3',
        'Master IV': 'VETERANI 4'
    }
    return translations.get(division_type, division_type.upper())

def auto_fit_columns(worksheet):
    """Auto-fit all column widths in the worksheet"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        
        # Find the first valid cell to get the column letter
        for cell in column:
            try:
                if hasattr(cell, 'column_letter') and cell.column_letter:
                    if column_letter is None:
                        column_letter = cell.column_letter
                    
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            except:
                pass
        
        # Apply width if we found a valid column letter
        if column_letter is not None:
            adjusted_width = min(max_length + 3, 30)
            if adjusted_width < 8:  # Minimum width
                adjusted_width = 8
            worksheet.column_dimensions[column_letter].width = adjusted_width

def sort_by_categories(df):
    """Sort dataframe by division order, then weight class, then place"""
    
    # Define division order
    division_order = {
        'Sub-Junior': 1,
        'Junior': 2, 
        'Open': 3,
        'Master I': 4,
        'Master II': 5,
        'Master III': 6,
        'Master IV': 7
    }
    
    # Add division order column (handle NaN values)
    df['DivisionOrder'] = df['Division'].fillna('Open').apply(lambda x: division_order.get(get_division_type(x), 3))
    
    # Convert WeightClassKg to numeric for proper sorting, handling superheavyweight classes
    def weight_sort_key(weight_class_str):
        """Convert weight class to sortable numeric value, handling + classes"""
        weight_str = str(weight_class_str)
        try:
            if '+' in weight_str:
                # For superheavyweight classes like "120+" or "84+", add 0.5 to sort after the base weight
                base_weight = float(weight_str.replace('+', ''))
                return base_weight + 0.5
            else:
                return float(weight_str)
        except ValueError:
            # For non-numeric values like "All Guest", return a very high number to sort at end
            return 9999.0
    
    df['WeightClassKg_num'] = df['WeightClassKg'].apply(weight_sort_key)
    
    # Convert Place to numeric for proper sorting
    df['Place_num'] = pd.to_numeric(df['Place'], errors='coerce')
    
    # Sort by division order, then weight class, then place
    df_sorted = df.sort_values(['DivisionOrder', 'WeightClassKg_num', 'Place_num'])
    
    # Remove helper columns
    df_sorted = df_sorted.drop(['DivisionOrder', 'WeightClassKg_num', 'Place_num'], axis=1)
    
    return df_sorted

def create_pretty_excel(equipment_filter: str = 'Raw', output_filename: str = None):
    """Create a beautifully formatted Excel file from processed powerlifting data.
    equipment_filter: 'Raw' | 'Equipped' | None (None = no filter)
    output_filename: optional explicit output filename
    """
    
    # Read the processed data
    df_full = pd.read_csv('powerlifting_results_processed.csv')
    
    # Provjeri da li postoji Equipment kolona
    has_equipment = 'Equipment' in df_full.columns
    
    # Apply equipment filtering za rezultate (samo ako je specificiran filter)
    # Za rang klubova koristimo puni dataset da vidimo Raw i Equipped odvojeno
    if equipment_filter in ('Raw', 'Equipped') and has_equipment:
        if equipment_filter == 'Raw':
            df = df_full[df_full['Equipment'] == 'Raw'].copy()
        else:  # Equipped
            df = df_full[df_full['Equipment'] == 'Equipped'].copy()
    else:
        df = df_full.copy()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles with modern color scheme
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')  # Modern navy blue
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='D9D9D9'),  # Light gray borders
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 1. Men's Powerlifting Sheet
    print("Kreiranje 'Muški Powerlifting' stranice...")
    men_sbd = df[(df['Sex'] == 'M') & (df['Event'] == 'SBD')].copy()
    men_sbd = sort_by_categories(men_sbd)
    
    ws_men_sbd = wb.create_sheet("Muški Powerlifting")
    create_formatted_sheet(ws_men_sbd, men_sbd, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 2. Women's Powerlifting Sheet
    print("Kreiranje 'Ženski Powerlifting' stranice...")
    women_sbd = df[(df['Sex'] == 'F') & (df['Event'] == 'SBD')].copy()
    women_sbd = sort_by_categories(women_sbd)
    
    ws_women_sbd = wb.create_sheet("Ženski Powerlifting")
    create_formatted_sheet(ws_women_sbd, women_sbd, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 3. Men's Bench Only Sheet
    print("Kreiranje 'Muški Potisak s klupe' stranice...")
    men_bench = df[(df['Sex'] == 'M') & (df['Event'] == 'B')].copy()
    men_bench = sort_by_categories(men_bench)
    
    # Remove squat and deadlift columns for bench only
    squat_deadlift_columns = ['Squat1Kg', 'Squat2Kg', 'Squat3Kg', 'Best3SquatKg', 
                             'Deadlift1Kg', 'Deadlift2Kg', 'Deadlift3Kg', 'Best3DeadliftKg']
    men_bench_filtered = men_bench.drop(columns=[col for col in squat_deadlift_columns if col in men_bench.columns])
    
    ws_men_bench = wb.create_sheet("Muški Potisak s klupe")
    create_formatted_sheet(ws_men_bench, men_bench_filtered, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 4. Women's Bench Only Sheet
    print("Kreiranje 'Ženski Potisak s klupe' stranice...")
    women_bench = df[(df['Sex'] == 'F') & (df['Event'] == 'B')].copy()
    women_bench = sort_by_categories(women_bench)
    
    # Remove squat and deadlift columns for bench only
    women_bench_filtered = women_bench.drop(columns=[col for col in squat_deadlift_columns if col in women_bench.columns])
    
    ws_women_bench = wb.create_sheet("Ženski Potisak s klupe")
    create_formatted_sheet(ws_women_bench, women_bench_filtered, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 5. Club Rankings Summary Sheet
    print("Kreiranje 'Rang Klubova' stranice...")
    ws_clubs = wb.create_sheet("Rang Klubova")
    # Uvijek koristi puni dataset (bez filtera) za rang klubova da vidimo Raw i Equipped odvojeno
    create_club_summary_sheet_with_equipment(ws_clubs, df_full, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # 6. Statistics Sheet
    print("Kreiranje 'Statistika' stranice...")
    ws_stats = wb.create_sheet("Statistika")
    create_statistics_sheet(ws_stats, df, header_font, header_fill, header_alignment, data_font, data_alignment, border)
    
    # Set the first sheet as active
    wb.active = wb["Muški Powerlifting"]
    
    # Save the workbook
    if output_filename:
        filename = output_filename
    else:
        if equipment_filter == 'Equipped':
            filename = "rezultati_equipped.xlsx"
        elif equipment_filter == 'Raw':
            filename = "rezultati.xlsx"
        else:
            filename = "rezultati.xlsx"
    wb.save(filename)
    print(f"\n[OK] Excel datoteka '{filename}' uspjesno kreirana!")
    print(f"Ukupno stranica: {len(wb.sheetnames)}")
    print(f"Ukupno zapisa: {len(df)}")
    print("Stranice: Muski Powerlifting, Zenski Powerlifting, Muski Potisak s klupe, Zenski Potisak s klupe, Rang Klubova, Statistika")
    
    return filename

def create_formatted_sheet(worksheet, data, header_font, header_fill, header_alignment, data_font, data_alignment, border):
    """Create a formatted sheet with separate tables for each division/weight class combination"""
    
    if len(data) == 0:
        return
    
    current_row = 1
    
    # Define division order (same as statistics)
    division_order = ['Sub-Junior', 'Junior', 'Open', 'Master I', 'Master II', 'Master III', 'Master IV']
    
    # Create a copy to avoid modifying original data
    data_copy = data.copy()
    
    # Create division type mapping (handle NaN values)
    data_copy['DivisionType'] = data_copy['Division'].fillna('Open').apply(get_division_type)
    
    # Create weight class sorting key that handles superheavyweight classes
    def weight_sort_key(weight_class_str):
        """Convert weight class to sortable numeric value, handling + classes"""
        weight_str = str(weight_class_str)
        try:
            if '+' in weight_str:
                # For superheavyweight classes like "120+" or "84+", add 0.5 to sort after the base weight
                base_weight = float(weight_str.replace('+', ''))
                return base_weight + 0.5
            else:
                return float(weight_str)
        except ValueError:
            # For non-numeric values like "All Guest", return a very high number to sort at end
            return 9999.0
    
    data_copy['WeightSortKey'] = data_copy['WeightClassKg'].apply(weight_sort_key)
    
    # Convert Place to numeric for proper sorting
    data_copy['PlaceNumeric'] = pd.to_numeric(data_copy['Place'], errors='coerce')
    
    # Sort data by division type (using custom order) then by weight class, then by numeric place
    division_order_map = {div: i for i, div in enumerate(division_order)}
    data_copy['DivisionOrder'] = data_copy['DivisionType'].map(division_order_map).fillna(999)  # Unknown divisions at end
    data_sorted = data_copy.sort_values(['DivisionOrder', 'WeightSortKey', 'PlaceNumeric'])
    
    # Get original column names (excluding helper columns, Division and WeightClassKg since they're shown in headers)
    original_columns = [col for col in data.columns if col not in ['Division', 'WeightClassKg', 'Sex']]
    
    # Get unique division/weight class combinations in sorted order
    unique_combinations = []
    seen = set()
    for _, row in data_sorted.iterrows():
        # Handle NaN values
        division_val = row['Division'] if pd.notna(row['Division']) else 'Open'
        weight_class_val = row['WeightClassKg'] if pd.notna(row['WeightClassKg']) else ''
        combo = (division_val, weight_class_val)
        if combo not in seen:
            unique_combinations.append(combo)
            seen.add(combo)
    
    # Process each division/weight class combination in sorted order
    current_division_type = None
    
    for division, weight_class in unique_combinations:
        # Filter data for this specific combination
        # Handle NaN values in comparison
        division_mask = data_sorted['Division'].fillna('Open') == division
        weight_mask = data_sorted['WeightClassKg'].fillna('') == weight_class
        group_data = data_sorted[division_mask & weight_mask]
        
        # Get division type for this combination
        division_type = get_division_type(division)
        
        # Add division type separator if this is a new division type
        if current_division_type != division_type:
            if current_division_type is not None:  # Not the first division
                current_row += 2  # Extra spacing between division types
            
            # Add division type header with different styling
            translated_division_type = translate_division_type(division_type)
            division_header = f"═══ {translated_division_type} KATEGORIJA ═══"
            cell = worksheet.cell(row=current_row, column=1, value=division_header)
            cell.font = Font(size=14, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0F2B47', end_color='0F2B47', fill_type='solid')  # Darker navy for division headers
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Merge cells across all columns for the division header
            if len(original_columns) > 1:
                worksheet.merge_cells(start_row=current_row, start_column=1, 
                                    end_row=current_row, end_column=len(original_columns))
            
            current_division_type = division_type
            current_row += 3  # Extra space after division header
        
        # Add category header (weight class specific, except for guests)
        translated_division = translate_division_name(division)
        division_str = str(division) if pd.notna(division) else ''
        if 'Guest' in division_str:
            category_title = translated_division  # No weight class for guests
        else:
            category_title = f"{translated_division} - {weight_class}kg"
        worksheet.cell(row=current_row, column=1, value=category_title).font = Font(size=12, bold=True)
        current_row += 2
        
        # Add column headers (translated to Croatian)
        translated_headers = translate_column_headers(original_columns)
        for col_idx, header in enumerate(translated_headers, 1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        # Add data rows for this category (excluding helper columns)
        for _, row_data in group_data.iterrows():
            # Determine if this row should have medal coloring
            place_value = str(row_data['Place']).strip()
            medal_fill = None
            
            if place_value == '1':
                medal_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Gold
            elif place_value == '2':
                medal_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # Silver
            elif place_value == '3':
                medal_fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')  # Bronze
            
            for col_idx, col_name in enumerate(original_columns, 1):
                value = row_data[col_name]
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Apply medal coloring if applicable
                if medal_fill is not None:
                    cell.fill = medal_fill
            
            current_row += 1
        
        # Add empty row between categories
        current_row += 1
    
    # Auto-adjust column widths
    auto_fit_columns(worksheet)


def create_club_summary_sheet_with_equipment(worksheet, df, header_font, header_fill, header_alignment, data_font, data_alignment, border):
    """Create club rankings summary sheet with separate Raw and Equipped rankings."""
    
    # Provjeri da li postoji Equipment kolona
    has_equipment = 'Equipment' in df.columns
    
    # Definiraj kategorije
    categories = [
        ((df['Sex'] == 'M') & (df['Event'] == 'SBD'), 'Muški Powerlifting'),
        ((df['Sex'] == 'F') & (df['Event'] == 'SBD'), 'Ženski Powerlifting'),
        ((df['Sex'] == 'M') & (df['Event'] == 'B'), 'Muški Potisak s klupe'),
        ((df['Sex'] == 'F') & (df['Event'] == 'B'), 'Ženski Potisak s klupe')
    ]
    
    current_row = 1
    
    for mask, category_name in categories:
        category_data = df[mask]
        if category_data.empty:
            continue
        
        # Dodaj naslov kategorije
        worksheet.cell(row=current_row, column=1, value=f"{category_name} Rang Klubova").font = Font(size=14, bold=True)
        current_row += 2
        
        # Ako ima Equipment kolonu, odvoji Raw i Equipped
        if has_equipment:
            # RAW rang (bez naslova - podrazumijeva se)
            raw_data = category_data[category_data['Equipment'] == 'Raw']
            
            # Provjeri da li ima Equipped podataka
            equipped_data = category_data[category_data['Equipment'] == 'Equipped']
            has_equipped = not equipped_data.empty
            
            if not raw_data.empty:
                # Headers (bez "RAW" naslova)
                headers = ['Mjesto', 'Klub', 'Bodovi']
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                current_row += 1
                
                # Compute top-5 per club by Points, then sum
                # Uzmi samo top-5 natjecatelja po klubu
                top5_per_club = raw_data.sort_values('Points', ascending=False).groupby('Club').head(5)
                club_points = top5_per_club.groupby('Club')['Points'].sum().reset_index()
                club_points = club_points.sort_values('Points', ascending=False).reset_index(drop=True)
                club_points['Place'] = range(1, len(club_points) + 1)
                
                for _, row in club_points.iterrows():
                    for col_idx, value in enumerate([row['Place'], row['Club'], round(row['Points'], 2)], 1):
                        cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = border
                        if row['Place'] == 1:
                            cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        elif row['Place'] == 2:
                            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                        elif row['Place'] == 3:
                            cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                    current_row += 1
                
                # Samo ako ima Equipped, dodaj razmak
                if has_equipped:
                    current_row += 2
            
            # EQUIPPED rang (samo ako postoji)
            if has_equipped:
                worksheet.cell(row=current_row, column=1, value="EQUIPPED").font = Font(size=12, bold=True, color='C65911')
                current_row += 1
                
                # Headers
                headers = ['Mjesto', 'Klub', 'Bodovi']
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                current_row += 1
                
                # Compute top-5 per club by Points, then sum
                # Uzmi samo top-5 natjecatelja po klubu
                top5_per_club = equipped_data.sort_values('Points', ascending=False).groupby('Club').head(5)
                club_points = top5_per_club.groupby('Club')['Points'].sum().reset_index()
                club_points = club_points.sort_values('Points', ascending=False).reset_index(drop=True)
                club_points['Place'] = range(1, len(club_points) + 1)
                
                for _, row in club_points.iterrows():
                    for col_idx, value in enumerate([row['Place'], row['Club'], round(row['Points'], 2)], 1):
                        cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = border
                        if row['Place'] == 1:
                            cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        elif row['Place'] == 2:
                            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                        elif row['Place'] == 3:
                            cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                    current_row += 1
                current_row += 2  # Space before next category
            else:
                # Ako nema Equipped, samo dodaj razmak
                current_row += 2
        else:
            # Ako nema Equipment kolonu, koristi standardni pristup
            # Headers
            headers = ['Mjesto', 'Klub', 'Bodovi']
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            current_row += 1
            
            # Compute top-5 per club by Points, then sum
            # Uzmi samo top-5 natjecatelja po klubu
            top5_per_club = category_data.sort_values('Points', ascending=False).groupby('Club').head(5)
            club_points = top5_per_club.groupby('Club')['Points'].sum().reset_index()
            club_points = club_points.sort_values('Points', ascending=False).reset_index(drop=True)
            club_points['Place'] = range(1, len(club_points) + 1)
            
            for _, row in club_points.iterrows():
                for col_idx, value in enumerate([row['Place'], row['Club'], round(row['Points'], 2)], 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    if row['Place'] == 1:
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                    elif row['Place'] == 2:
                        cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                    elif row['Place'] == 3:
                        cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                current_row += 1
            current_row += 2  # Space before next category
    
    auto_fit_columns(worksheet)


def create_statistics_sheet(worksheet, df, header_font, header_fill, header_alignment, data_font, data_alignment, border):
    """Create statistics summary sheet"""
    
    # Title
    worksheet.cell(row=1, column=1, value="Statistika Natjecanja").font = Font(size=16, bold=True)
    
    current_row = 3
    
    # General statistics
    stats = [
        ("Ukupno Natjecatelja", len(df)),
        ("Muških Natjecatelja", len(df[df['Sex'] == 'M'])),
        ("Ženskih Natjecatelja", len(df[df['Sex'] == 'F'])),
        ("Powerlifting", len(df[df['Event'] == 'SBD'])),
        ("Potisak s klupe", len(df[df['Event'] == 'B'])),
        ("Ukupno Klubova", df['Club'].nunique()),
        ("Prosjek GL Bodova", f"{df['Points'].mean():.2f}"),
        ("Najbolji GL Bodovi", f"{df['Points'].max():.2f}")
    ]
    
    # Add statistics
    for stat_name, stat_value in stats:
        worksheet.cell(row=current_row, column=1, value=stat_name).font = Font(bold=True)
        stat_cell = worksheet.cell(row=current_row, column=2, value=stat_value)
        stat_cell.alignment = data_alignment  # Apply consistent alignment
        current_row += 1
    
    current_row += 2
    
    # Get unique divisions and sort them
    division_order = ['Sub-Junior', 'Junior', 'Open', 'Master I', 'Master II', 'Master III', 'Master IV']
    
    # Group by division type (handle NaN values)
    df['DivisionType'] = df['Division'].fillna('Open').apply(get_division_type)
    
    def create_top_5_section(title, data, current_row):
        """Helper function to create a top 5 section with optional Raw/Equipped split"""
        if len(data) == 0:
            return current_row
        
        # Check if we should split by equipment
        has_equipment = 'Equipment' in data.columns
        has_equipped = has_equipment and (data['Equipment'] == 'Equipped').any()
        
        if has_equipment and has_equipped:
            # RAW Top 5 (bez naslova - podrazumijeva se)
            raw_data = data[data['Equipment'] == 'Raw']
            if len(raw_data) > 0:
                worksheet.cell(row=current_row, column=1, value=title).font = Font(size=12, bold=True)
                current_row += 2
                
                # Headers
                headers = ['Rang', 'Ime', 'Klub', 'Ukupno (kg)', 'GL Bodovi']
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                current_row += 1
                
                # Top 5 in category
                top_5 = raw_data.nlargest(5, 'Points')
                for rank, (_, performer) in enumerate(top_5.iterrows(), 1):
                    values = [rank, performer['Name'], performer['Club'],
                             performer['TotalKg'], f"{performer['Points']:.2f}"]
                    
                    for col_idx, value in enumerate(values, 1):
                        cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = border
                        
                        # Medal colors
                        if rank == 1:
                            cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        elif rank == 2:
                            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                        elif rank == 3:
                            cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                    
                    current_row += 1
                
                current_row += 2  # Space after section
            
            # EQUIPPED Top 5 (samo ako postoji)
            equipped_data = data[data['Equipment'] == 'Equipped']
            if len(equipped_data) > 0:
                worksheet.cell(row=current_row, column=1, value=f"{title} - EQUIPPED").font = Font(size=12, bold=True, color='C65911')
                current_row += 2
                
                # Headers
                headers = ['Rang', 'Ime', 'Klub', 'Ukupno (kg)', 'GL Bodovi']
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                current_row += 1
                
                # Top 5 in category
                top_5 = equipped_data.nlargest(5, 'Points')
                for rank, (_, performer) in enumerate(top_5.iterrows(), 1):
                    values = [rank, performer['Name'], performer['Club'],
                             performer['TotalKg'], f"{performer['Points']:.2f}"]
                    
                    for col_idx, value in enumerate(values, 1):
                        cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = border
                        
                        # Medal colors
                        if rank == 1:
                            cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        elif rank == 2:
                            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                        elif rank == 3:
                            cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                    
                    current_row += 1
                
                current_row += 2  # Space after section
        else:
            # Standard approach (no equipment split)
            worksheet.cell(row=current_row, column=1, value=title).font = Font(size=12, bold=True)
            current_row += 2
            
            # Headers
            headers = ['Rang', 'Ime', 'Klub', 'Ukupno (kg)', 'GL Bodovi']
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            current_row += 1
            
            # Top 5 in category
            top_5 = data.nlargest(5, 'Points')
            for rank, (_, performer) in enumerate(top_5.iterrows(), 1):
                values = [rank, performer['Name'], performer['Club'],
                         performer['TotalKg'], f"{performer['Points']:.2f}"]
                
                for col_idx, value in enumerate(values, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # Medal colors
                    if rank == 1:
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                    elif rank == 2:
                        cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                    elif rank == 3:
                        cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
                
                current_row += 1
            
            current_row += 2  # Space after section
        
        return current_row
    
    # 1. MALE POWERLIFTING SECTION
    male_powerlifting = df[(df['Sex'] == 'M') & (df['Event'] == 'SBD')]
    
    # Overall Male Powerlifting
    current_row = create_top_5_section("Top 5 Muški Powerlifting", male_powerlifting, current_row)
    
    # Male Powerlifting by Division
    for division_type in division_order:
        division_data = male_powerlifting[male_powerlifting['DivisionType'] == division_type]
        if len(division_data) > 0:
            translated_division_type = translate_division_type(division_type).title()
            current_row = create_top_5_section(f"Top 5 Muški {translated_division_type} Powerlifting", division_data, current_row)
    
    # 2. FEMALE POWERLIFTING SECTION
    female_powerlifting = df[(df['Sex'] == 'F') & (df['Event'] == 'SBD')]
    
    # Overall Female Powerlifting
    current_row = create_top_5_section("Top 5 Ženski Powerlifting", female_powerlifting, current_row)
    
    # Female Powerlifting by Division
    for division_type in division_order:
        division_data = female_powerlifting[female_powerlifting['DivisionType'] == division_type]
        if len(division_data) > 0:
            translated_division_type = translate_division_type(division_type).title()
            current_row = create_top_5_section(f"Top 5 Ženski {translated_division_type} Powerlifting", division_data, current_row)
    
    # 3. MALE BENCH ONLY SECTION
    male_bench = df[(df['Sex'] == 'M') & (df['Event'] == 'B')]
    
    # Overall Male Bench Only
    current_row = create_top_5_section("Top 5 Muški Potisak s klupe", male_bench, current_row)
    
    # Male Bench Only by Division
    for division_type in division_order:
        division_data = male_bench[male_bench['DivisionType'] == division_type]
        if len(division_data) > 0:
            translated_division_type = translate_division_type(division_type).title()
            current_row = create_top_5_section(f"Top 5 Muški {translated_division_type} Potisak s klupe", division_data, current_row)
    
    # 4. FEMALE BENCH ONLY SECTION
    female_bench = df[(df['Sex'] == 'F') & (df['Event'] == 'B')]
    
    # Overall Female Bench Only
    current_row = create_top_5_section("Top 5 Ženski Potisak s klupe", female_bench, current_row)
    
    # Female Bench Only by Division
    for division_type in division_order:
        division_data = female_bench[female_bench['DivisionType'] == division_type]
        if len(division_data) > 0:
            translated_division_type = translate_division_type(division_type).title()
            current_row = create_top_5_section(f"Top 5 Ženski {translated_division_type} Potisak s klupe", division_data, current_row)
    
    # Auto-adjust column widths
    auto_fit_columns(worksheet)

if __name__ == "__main__":
    try:
        create_pretty_excel('Raw')
    except Exception as e:
        print(f"Greška pri kreiranju Excel datoteke: {e}")
        print("Provjerite je li instaliran openpyxl: pip install openpyxl")